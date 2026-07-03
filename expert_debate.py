from __future__ import annotations

import json
import re
import urllib.error
import urllib.request
from datetime import date
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from .ai_advisor import extract_advice_text, load_ai_config
from .config import ROOT_DIR
from .db import connect, json_dumps, new_id, now_iso, row_to_dict, rows_to_dicts
from .engine import data_health, latest_report, portfolio_state
from .universe import universe_rows, universe_summary


REPORT_DIR = ROOT_DIR / "output" / "commercial_reports"
PDF_DIR = ROOT_DIR / "output" / "pdf"
EXCEL_DIR = ROOT_DIR / "output" / "excel"

COMMON_GUARDRAILS = """
You are working for SimVest, a simulation-only investment research platform.
Hard constraints:
- Simulation only. Do not imply that real broker orders have been placed.
- No leverage, no short selling, no derivatives speculation.
- Only public-market, ordinary-account, low-threshold products are allowed.
- Every trade action must wait for human confirmation before it is recorded.
- Benchmark comparison: CSI 300 and S&P 500, natural-year net return after costs.
- Maximum preferred drawdown is 20%; single-stock and single-asset caps must be respected.
- Distinguish facts, model inferences, confidence and missing data.
- Output confidence in 5% increments.
"""

EXPERT_ROLES = [
    {
        "id": "public_data_expert",
        "name_cn": "公开数据与信源可靠性专家",
        "name_en": "Public Data and Source Reliability Expert",
        "prompt": f"""
{COMMON_GUARDRAILS}
Role: Public Data and Source Reliability Expert.

Your job is to audit and interpret the public data layer before any investment
decision is discussed. Focus on:
1. Coverage: A shares, Hong Kong stocks, funds/ETFs, bonds, FX, futures proxies,
   gold, energy, macro, geopolitics, technology and company events.
2. Freshness: quote dates, source run dates, event dates and stale-data risks.
3. Reliability: separate primary facts from vendor-derived fields and model labels.
4. Market signal extraction: describe which public data points are actually strong
   enough to support today's simulated decision.
5. Data gaps: list what must not be over-interpreted.

Return bilingual notes with these sections:
- 中文结论
- English conclusion
- Evidence ledger
- Data gaps and confidence
""".strip(),
    },
    {
        "id": "portfolio_valuation_expert",
        "name_cn": "组合估值与风险预算专家",
        "name_en": "Portfolio Valuation and Risk Budget Expert",
        "prompt": f"""
{COMMON_GUARDRAILS}
Role: Portfolio Valuation and Risk Budget Expert.

Your job is to mark the portfolio to the latest available public prices and judge
whether the proposed simulated actions are consistent with risk limits. Focus on:
1. Total value, cash, unrealized P/L, invested-cost P/L and position weights.
2. Exposure by asset bucket and concentration risks.
3. Benchmark comparison against CSI 300 and S&P 500.
4. Whether to buy, hold, trim, or wait; include target weights, reference prices,
   stop-loss, take-profit and holding horizon when the context supports it.
5. Transaction cost, slippage and human-confirmation constraints.

Return bilingual notes with these sections:
- 中文结论
- English conclusion
- Position-level risk comments
- Disagreements or vetoes
""".strip(),
    },
    {
        "id": "event_causality_expert",
        "name_cn": "事件因果链与跨资产传导专家",
        "name_en": "Event Causality and Cross-Asset Transmission Expert",
        "prompt": f"""
{COMMON_GUARDRAILS}
Role: Event Causality and Cross-Asset Transmission Expert.

Your job is to turn today's market, macro, technology, company and geopolitical
information into one coherent causal chain. Focus on:
1. What changed today or recently.
2. The route from event -> liquidity/rates/risk appetite/earnings/FX/commodities
   -> sectors -> portfolio actions.
3. Alternative explanations and what evidence would falsify the chain.
4. Cross-asset consequences for stocks, bonds, FX, gold and energy.
5. Whether the causal chain is strong enough to support new simulated trades.

Return bilingual notes with these sections:
- 中文因果链
- English causal chain
- Asset-class implications
- Confidence and invalidation signals
""".strip(),
    },
]

CHAIR_PROMPT = f"""
{COMMON_GUARDRAILS}
Role: Investment Committee Chair and Commercial Report Editor.

You will receive three expert notes plus the SimVest context. Moderate a concise
debate, resolve disagreements, and produce a commercial-ready daily report.

The final report must be bilingual and must use this exact high-level structure:

# SimVest Expert Committee Daily Report
# SimVest 专家委员会每日报告

## 中文报告
### 1. 执行摘要
### 2. 今日一条完整主线
### 3. 三位专家辩论后的共识与分歧
### 4. 组合估值与盈亏
### 5. 今日模拟投资决策
### 6. 风险、失效条件与不做什么
### 7. 数据质量与后续待补强

## English Report
### 1. Executive Summary
### 2. One Coherent Market Line
### 3. Committee Consensus and Disagreements
### 4. Portfolio Valuation and P/L
### 5. Today's Simulated Investment Decisions
### 6. Risks, Invalidation and What Not To Do
### 7. Data Quality and Next Improvements

Requirements:
- Keep Chinese and English in separate sections so the PDF can apply SimSun to
  Chinese paragraphs and Times New Roman to English paragraphs.
- Be specific on symbols, product names, actions, target/current weights,
  reference price, stop-loss, take-profit, holding period and confidence when
  data is available.
- If data is insufficient for a new trade, explicitly say HOLD/WAIT and why.
- Never promise excess return over the benchmarks; phrase it as a simulation goal.
- End with a simulation-only disclaimer.
""".strip()


def expert_prompts() -> dict[str, Any]:
    return {
        "experts": EXPERT_ROLES,
        "chair": {
            "id": "investment_committee_chair",
            "name_cn": "投委会主席与商业报告编辑",
            "name_en": "Investment Committee Chair and Commercial Report Editor",
            "prompt": CHAIR_PROMPT,
        },
    }


def build_debate_context() -> dict[str, Any]:
    report = latest_report()
    health = compact_health(data_health())
    universe_context = {
        "summary": universe_summary(),
        "top_turnover": compact_universe_rows(
            universe_rows({"limit": ["24"], "sort": ["turnover"], "direction": ["desc"]})["rows"]
        ),
        "top_gainers": compact_universe_rows(
            universe_rows({"limit": ["16"], "sort": ["change"], "direction": ["desc"]})["rows"]
        ),
        "top_losers": compact_universe_rows(
            universe_rows({"limit": ["16"], "sort": ["change"], "direction": ["asc"]})["rows"]
        ),
    }
    with connect() as conn:
        portfolio = portfolio_state(conn)
        assets = rows_to_dicts(
            conn.execute(
                """
                SELECT symbol, name_cn, name_en, asset_class, product_type, bucket, market,
                       region, currency, price, prev_close, day_change_pct, ytd_return_pct,
                       volatility_20d, risk_bucket, tradable
                FROM assets
                ORDER BY asset_class, market, symbol
                """
            ).fetchall()
        )
        benchmarks = rows_to_dicts(conn.execute("SELECT * FROM benchmarks ORDER BY id").fetchall())
        events = rows_to_dicts(
            conn.execute(
                """
                SELECT created_at, title_cn, title_en, source_type, region, category,
                       severity, sentiment, confidence, link, notes_cn, notes_en
                FROM events
                ORDER BY created_at DESC
                LIMIT 45
                """
            ).fetchall()
        )
        source_documents = rows_to_dicts(
            conn.execute(
                """
                SELECT fetched_at, source, source_url, title_cn, title_en, published_at,
                       region, category, severity, sentiment, confidence
                FROM raw_documents
                ORDER BY fetched_at DESC
                LIMIT 30
                """
            ).fetchall()
        )
        snapshots = rows_to_dicts(
            conn.execute(
                """
                SELECT as_of, headline_cn, headline_en, csi300_change, sp500_change,
                       hsi_change, usdcnh_change, gold_change, oil_change,
                       policy_signal, geopolitics_signal, liquidity_signal, source_quality
                FROM market_snapshots
                ORDER BY rowid DESC
                LIMIT 5
                """
            ).fetchall()
        )
        performance = rows_to_dicts(
            conn.execute(
                """
                SELECT *
                FROM portfolio_daily
                ORDER BY as_of DESC
                LIMIT 20
                """
            ).fetchall()
        )
    return {
        "as_of": date.today().isoformat(),
        "simulation_scope": {
            "initial_capital_cny": 150000,
            "markets": ["China A-share", "Hong Kong"],
            "benchmarks": ["CSI 300", "S&P 500"],
            "constraints": [
                "simulation only",
                "manual confirmation required",
                "no leverage",
                "no short selling",
                "public-market low-threshold products only",
            ],
        },
        "portfolio": compact_portfolio(portfolio),
        "benchmarks": compact_rows(benchmarks),
        "assets": compact_rows(assets),
        "market_universe": universe_context,
        "events": compact_rows(events),
        "source_documents": compact_rows(source_documents),
        "market_snapshots": compact_rows(snapshots),
        "performance_recent": list(reversed(compact_rows(performance))),
        "latest_report": compact_latest_report(report),
        "data_health": health,
    }


def run_expert_debate() -> dict[str, Any]:
    config = load_ai_config()
    context = build_debate_context()
    report_id = new_id("expert")
    created_at = now_iso()
    prompts = expert_prompts()
    expert_outputs: list[dict[str, str]] = []
    final_report = ""
    status = "ok"
    error = ""

    if not config["api_key"]:
        status = "missing_key"
        error = "Missing SIMVEST_LLM_API_KEY or data/secrets.json api_key"
        final_report = fallback_report(context, expert_outputs, status)
    else:
        try:
            for role in EXPERT_ROLES:
                output = call_llm(
                    config,
                    [
                        {"role": "system", "content": role["prompt"]},
                        {
                            "role": "user",
                            "content": "Analyze this SimVest context and return your expert notes.\n\n"
                            + json.dumps(context, ensure_ascii=False),
                        },
                    ],
                    max_tokens=1800,
                    temperature=0.15,
                )
                expert_outputs.append(
                    {
                        "id": role["id"],
                        "name_cn": role["name_cn"],
                        "name_en": role["name_en"],
                        "output": output,
                    }
                )
            chair_context = {
                "simvest_context": context,
                "expert_outputs": expert_outputs,
            }
            final_report = call_llm(
                config,
                [
                    {"role": "system", "content": CHAIR_PROMPT},
                    {
                        "role": "user",
                        "content": "Resolve the expert debate and write the final bilingual report.\n\n"
                        + json.dumps(chair_context, ensure_ascii=False),
                    },
                ],
                max_tokens=4200,
                temperature=0.18,
            )
        except Exception as exc:  # pragma: no cover - external LLM boundary
            status = "error"
            error = str(exc)
            final_report = fallback_report(context, expert_outputs, status)

    markdown_path = save_markdown(report_id, final_report)
    pdf_path = build_pdf_report(report_id, final_report)
    excel_path = build_excel_report(report_id, final_report, expert_outputs, prompts, context)
    saved = save_expert_report(
        report_id=report_id,
        created_at=created_at,
        config=config,
        status=status,
        prompts=prompts,
        expert_outputs=expert_outputs,
        final_report=final_report,
        context=context,
        markdown_path=markdown_path,
        pdf_path=pdf_path,
        excel_path=excel_path,
        error=error,
    )
    return saved


def call_llm(
    config: dict[str, str],
    messages: list[dict[str, str]],
    max_tokens: int,
    temperature: float,
) -> str:
    payload = {
        "model": config["model"],
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    req = urllib.request.Request(
        config["base_url"],
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {config['api_key']}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=90) as response:
            raw = json.loads(response.read().decode("utf-8"))
        return extract_advice_text(raw).strip()
    except urllib.error.HTTPError as exc:
        error_body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"LLM HTTP {exc.code}: {error_body[:1200]}") from exc


def save_expert_report(
    report_id: str,
    created_at: str,
    config: dict[str, str],
    status: str,
    prompts: dict[str, Any],
    expert_outputs: list[dict[str, str]],
    final_report: str,
    context: dict[str, Any],
    markdown_path: Path,
    pdf_path: Path,
    excel_path: Path | None,
    error: str,
) -> dict[str, Any]:
    with connect() as conn:
        conn.execute(
            """
            INSERT INTO expert_debate_reports (
                id, created_at, as_of, provider, model, status, prompts_json,
                expert_outputs_json, final_report_md, context_json,
                markdown_path, pdf_path, excel_path, error
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                report_id,
                created_at,
                context["as_of"],
                config.get("provider", "openai-compatible"),
                config.get("model", ""),
                status,
                json_dumps(prompts),
                json_dumps(expert_outputs),
                final_report,
                json_dumps(context),
                str(markdown_path),
                str(pdf_path),
                str(excel_path) if excel_path else None,
                error,
            ),
        )
    return {
        "id": report_id,
        "created_at": created_at,
        "as_of": context["as_of"],
        "provider": config.get("provider", "openai-compatible"),
        "model": config.get("model", ""),
        "status": status,
        "prompts": prompts,
        "expert_outputs": expert_outputs,
        "final_report_md": final_report,
        "markdown_path": str(markdown_path),
        "pdf_path": str(pdf_path),
        "excel_path": str(excel_path) if excel_path else None,
        "error": error,
    }


def latest_expert_debate() -> dict[str, Any] | None:
    with connect() as conn:
        row = conn.execute(
            """
            SELECT id, created_at, as_of, provider, model, status, prompts_json,
                   expert_outputs_json, final_report_md, markdown_path, pdf_path,
                   excel_path, error
            FROM expert_debate_reports
            ORDER BY created_at DESC
            LIMIT 1
            """
        ).fetchone()
    data = row_to_dict(row)
    if not data:
        return None
    data["prompts"] = json.loads(data.pop("prompts_json") or "{}")
    data["expert_outputs"] = json.loads(data.pop("expert_outputs_json") or "[]")
    return data


def save_markdown(report_id: str, markdown: str) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = REPORT_DIR / f"{report_id}.md"
    path.write_text(markdown.strip() + "\n", encoding="utf-8")
    return path


def build_pdf_report(report_id: str, markdown: str) -> Path:
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    register_fonts()
    path = PDF_DIR / f"{report_id}.pdf"
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=15 * mm,
        bottomMargin=16 * mm,
        title="SimVest Expert Committee Daily Report",
        author="SimVest",
    )
    story = markdown_to_flowables(markdown)
    doc.build(story)
    return path


def register_fonts() -> None:
    fonts = {
        "SimSun": "C:/Windows/Fonts/simsun.ttc",
        "TimesNewRoman": "C:/Windows/Fonts/times.ttf",
        "TimesNewRoman-Bold": "C:/Windows/Fonts/timesbd.ttf",
    }
    for name, filename in fonts.items():
        if name not in pdfmetrics.getRegisteredFontNames() and Path(filename).exists():
            pdfmetrics.registerFont(TTFont(name, filename))


def markdown_to_flowables(markdown: str) -> list[Any]:
    base = getSampleStyleSheet()
    styles = {
        "title_cn": ParagraphStyle(
            "title_cn",
            parent=base["Title"],
            fontName="SimSun",
            fontSize=18,
            leading=23,
            spaceAfter=8,
            textColor=colors.HexColor("#182230"),
            wordWrap="CJK",
        ),
        "title_en": ParagraphStyle(
            "title_en",
            parent=base["Title"],
            fontName="TimesNewRoman-Bold",
            fontSize=18,
            leading=23,
            spaceAfter=8,
            textColor=colors.HexColor("#182230"),
        ),
        "heading_cn": ParagraphStyle(
            "heading_cn",
            parent=base["Heading2"],
            fontName="SimSun",
            fontSize=13,
            leading=17,
            spaceBefore=8,
            spaceAfter=4,
            textColor=colors.HexColor("#16665a"),
            wordWrap="CJK",
        ),
        "heading_en": ParagraphStyle(
            "heading_en",
            parent=base["Heading2"],
            fontName="TimesNewRoman-Bold",
            fontSize=13,
            leading=17,
            spaceBefore=8,
            spaceAfter=4,
            textColor=colors.HexColor("#16665a"),
        ),
        "body_cn": ParagraphStyle(
            "body_cn",
            parent=base["BodyText"],
            fontName="SimSun",
            fontSize=10,
            leading=14,
            spaceAfter=4,
            wordWrap="CJK",
        ),
        "body_en": ParagraphStyle(
            "body_en",
            parent=base["BodyText"],
            fontName="TimesNewRoman",
            fontSize=10,
            leading=14,
            spaceAfter=4,
        ),
        "small_cn": ParagraphStyle(
            "small_cn",
            parent=base["BodyText"],
            fontName="SimSun",
            fontSize=8.5,
            leading=12,
            spaceAfter=3,
            wordWrap="CJK",
        ),
        "small_en": ParagraphStyle(
            "small_en",
            parent=base["BodyText"],
            fontName="TimesNewRoman",
            fontSize=8.5,
            leading=12,
            spaceAfter=3,
        ),
    }
    story: list[Any] = []
    for raw_line in markdown.strip().splitlines():
        line = raw_line.strip()
        if not line:
            story.append(Spacer(1, 4))
            continue
        level = 0
        while level < len(line) and line[level] == "#":
            level += 1
        if level and len(line) > level and line[level] == " ":
            text = clean_markdown(line[level + 1 :])
            style = styles["title_cn" if contains_cjk(text) else "title_en"] if level == 1 else styles[
                "heading_cn" if contains_cjk(text) else "heading_en"
            ]
            story.append(Paragraph(escape(text), style))
            continue
        if line.startswith(("- ", "* ")):
            text = "- " + clean_markdown(line[2:])
            style = styles["small_cn" if contains_cjk(text) else "small_en"]
            story.append(Paragraph(escape(text), style))
            continue
        if line.startswith("|"):
            text = clean_markdown(line.replace("|", " | "))
            style = styles["small_cn" if contains_cjk(text) else "small_en"]
            story.append(Paragraph(escape(text), style))
            continue
        text = clean_markdown(line)
        style = styles["body_cn" if contains_cjk(text) else "body_en"]
        story.append(Paragraph(escape(text), style))
    return story


def build_excel_report(
    report_id: str,
    final_report: str,
    expert_outputs: list[dict[str, str]],
    prompts: dict[str, Any],
    context: dict[str, Any],
) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return None

    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    path = EXCEL_DIR / f"{report_id}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    write_sheet_rows(
        ws,
        [["SimVest Expert Committee Daily Report"], ["As of", context["as_of"]], [""]],
        Font,
        Alignment,
        PatternFill,
    )
    for line in final_report.splitlines():
        ws.append([line])
    style_sheet(ws, Font, Alignment, PatternFill)

    experts = wb.create_sheet("Expert Notes")
    experts.append(["Expert", "Chinese Name", "English Name", "Output"])
    for item in expert_outputs:
        experts.append([item.get("id"), item.get("name_cn"), item.get("name_en"), item.get("output")])
    style_sheet(experts, Font, Alignment, PatternFill)

    prompt_sheet = wb.create_sheet("Prompts")
    prompt_sheet.append(["Role", "Prompt"])
    for expert in prompts["experts"]:
        prompt_sheet.append([expert["id"], expert["prompt"]])
    prompt_sheet.append([prompts["chair"]["id"], prompts["chair"]["prompt"]])
    style_sheet(prompt_sheet, Font, Alignment, PatternFill)

    portfolio = wb.create_sheet("Portfolio")
    portfolio.append(["Metric", "Value"])
    for key, value in context.get("portfolio", {}).items():
        if key != "positions":
            portfolio.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    portfolio.append([])
    portfolio.append(["Symbol", "Name", "Asset Class", "Market Value", "Weight", "P/L", "P/L %"])
    for position in context.get("portfolio", {}).get("positions", []):
        portfolio.append(
            [
                position.get("symbol"),
                position.get("name_cn") or position.get("name_en"),
                position.get("asset_class"),
                position.get("market_value"),
                position.get("weight"),
                position.get("unrealized_pnl"),
                position.get("unrealized_pnl_pct"),
            ]
        )
    style_sheet(portfolio, Font, Alignment, PatternFill)
    wb.save(path)
    return path


def write_sheet_rows(ws: Any, rows: list[list[Any]], Font: Any, Alignment: Any, PatternFill: Any) -> None:
    for row in rows:
        ws.append(row)
    style_sheet(ws, Font, Alignment, PatternFill)


def style_sheet(ws: Any, Font: Any, Alignment: Any, PatternFill: Any) -> None:
    header_fill = PatternFill("solid", fgColor="E8F4F0")
    for row in ws.iter_rows():
        for cell in row:
            text = str(cell.value or "")
            cell.value = clean_cell_value(cell.value)
            cell.font = Font(name="SimSun" if contains_cjk(text) else "Times New Roman", size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row and row[0].row == 1:
            for cell in row:
                cell.fill = header_fill
                cell.font = Font(
                    name="SimSun" if contains_cjk(str(cell.value or "")) else "Times New Roman",
                    bold=True,
                    size=12,
                )
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        width = min(58, max(12, max(len(str(cell.value or "")) for cell in column_cells[:80]) + 2))
        ws.column_dimensions[column_letter].width = width


def fallback_report(context: dict[str, Any], expert_outputs: list[dict[str, str]], status: str) -> str:
    report = context.get("latest_report") or {}
    portfolio = context.get("portfolio") or {}
    decisions = report.get("decisions") or []
    decision_lines = []
    for decision in decisions[:8]:
        decision_lines.append(
            "- {symbol}: {action}, target {target_weight:.1%}, amount CNY {amount_cny:,.0f}, confidence {confidence}%.".format(
                symbol=decision.get("symbol", "--"),
                action=decision.get("action", "HOLD"),
                target_weight=float(decision.get("target_weight") or 0),
                amount_cny=float(decision.get("amount_cny") or 0),
                confidence=decision.get("confidence", "--"),
            )
        )
    if not decision_lines:
        decision_lines = ["- HOLD/WAIT: insufficient validated data for a new simulated action."]
    return f"""
# SimVest Expert Committee Daily Report
# SimVest 专家委员会每日报告

## 中文报告
### 1. 执行摘要
本报告由本地规则兜底生成，状态为 {status}。当前组合总市值约为 CNY {portfolio.get('total_value', 0):,.0f}，总盈亏约为 CNY {portfolio.get('total_pnl', 0):,.0f}，现金占比约为 {float(portfolio.get('cash_weight') or 0):.1%}。
### 2. 今日一条完整主线
最近一次系统报告主线为：{report.get('thesis_cn') or '数据不足，等待公开数据同步和专家模型输出。'}
### 3. 三位专家辩论后的共识与分歧
专家调用未完整完成，已保留已生成意见 {len(expert_outputs)} 条。当前不扩大风险暴露，以等待完整模型复核为主。
### 4. 组合估值与盈亏
组合净值、现金和持仓盈亏已按系统内最新公开价格估算；若行情源延迟，应以“观察/等待”为主。
### 5. 今日模拟投资决策
{chr(10).join(decision_lines)}
### 6. 风险、失效条件与不做什么
不做空、不使用杠杆、不自动下真实订单；若数据健康度下降、事件链无法验证或组合回撤接近预警线，暂停新增模拟交易。
### 7. 数据质量与后续待补强
需要继续补强公告、财务、行业比较和更长历史回测。

## English Report
### 1. Executive Summary
This report was generated by the local fallback engine with status {status}. Portfolio value is about CNY {portfolio.get('total_value', 0):,.0f}, total P/L is about CNY {portfolio.get('total_pnl', 0):,.0f}, and cash weight is about {float(portfolio.get('cash_weight') or 0):.1%}.
### 2. One Coherent Market Line
The latest system thesis is: {report.get('thesis_en') or report.get('thesis_cn') or 'insufficient data; wait for public data sync and model review.'}
### 3. Committee Consensus and Disagreements
The expert debate was not fully completed. {len(expert_outputs)} expert note(s) were retained. The fallback stance is to avoid increasing risk until the complete model review is available.
### 4. Portfolio Valuation and P/L
Portfolio value, cash and unrealized P/L are marked to the latest public prices available inside SimVest.
### 5. Today's Simulated Investment Decisions
{chr(10).join(decision_lines)}
### 6. Risks, Invalidation and What Not To Do
No short selling, no leverage and no real broker order. Pause new simulated trades if data health worsens, the causal chain cannot be verified, or drawdown approaches the warning band.
### 7. Data Quality and Next Improvements
Next improvements should add announcements, financial statements, sector comparison and longer backtests.

Simulation-only disclaimer: this report is for research and simulated investment records only. It is not real investment advice and does not guarantee returns over any benchmark.
""".strip()


def compact_portfolio(portfolio: dict[str, Any]) -> dict[str, Any]:
    return {
        "cash": portfolio.get("cash"),
        "cash_weight": portfolio.get("cash_weight"),
        "positions_value": portfolio.get("positions_value"),
        "total_value": portfolio.get("total_value"),
        "initial_capital": portfolio.get("initial_capital"),
        "total_pnl": portfolio.get("total_pnl"),
        "total_pnl_pct": portfolio.get("total_pnl_pct"),
        "invested_cost": portfolio.get("invested_cost"),
        "unrealized_pnl": portfolio.get("unrealized_pnl"),
        "unrealized_pnl_pct": portfolio.get("unrealized_pnl_pct"),
        "exposures": portfolio.get("exposures"),
        "positions": compact_rows(portfolio.get("positions", []), max_items=30),
    }


def compact_health(health: dict[str, Any]) -> dict[str, Any]:
    return {
        "status": health.get("status"),
        "status_cn": health.get("status_cn"),
        "score": health.get("score"),
        "quote_age_days": health.get("quote_age_days"),
        "document_age_days": health.get("document_age_days"),
        "universe_age_days": health.get("universe_age_days"),
        "coverage": health.get("coverage", []),
        "market_universe": health.get("market_universe"),
        "latest_runs": compact_rows(health.get("latest_runs", []), max_items=20),
    }


def compact_latest_report(report: dict[str, Any] | None) -> dict[str, Any] | None:
    if not report:
        return None
    return {
        "id": report.get("id"),
        "as_of": report.get("as_of"),
        "title_cn": report.get("title_cn"),
        "title_en": report.get("title_en"),
        "thesis_cn": report.get("thesis_cn"),
        "thesis_en": report.get("thesis_en"),
        "chain_cn": report.get("chain_cn"),
        "chain_en": report.get("chain_en"),
        "risk_level": report.get("risk_level"),
        "regime": report.get("regime"),
        "confidence": report.get("confidence"),
        "decisions": compact_rows(report.get("decisions", []), max_items=20),
    }


def compact_universe_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    keys = [
        "symbol",
        "code",
        "name_cn",
        "name_en",
        "market",
        "asset_class",
        "product_type",
        "sector",
        "board",
        "currency",
        "price",
        "prev_close",
        "day_change_pct",
        "turnover",
        "market_cap",
        "pe_ttm",
        "pb",
        "source",
        "updated_at",
    ]
    return [{key: item.get(key) for key in keys} for item in rows]


def compact_rows(rows: list[dict[str, Any]], max_items: int = 60) -> list[dict[str, Any]]:
    compacted = []
    for row in rows[:max_items]:
        item = {}
        for key, value in row.items():
            if key.endswith("_json") or key in {"payload_json", "raw_json", "context_json"}:
                continue
            if isinstance(value, str) and len(value) > 800:
                item[key] = value[:800] + "..."
            else:
                item[key] = value
        compacted.append(item)
    return compacted


def clean_markdown(text: str) -> str:
    text = text.replace("**", "").replace("__", "")
    text = text.replace("`", "")
    return re.sub(r"\s+", " ", text).strip()


def clean_cell_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", " ", value)
    return text[:32767]


def contains_cjk(text: str) -> bool:
    return any("\u4e00" <= char <= "\u9fff" for char in text)
